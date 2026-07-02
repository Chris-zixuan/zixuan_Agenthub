"""初始化脚本：生成 template.xlsx 工作文件，含分类映射和账户映射 sheet。

用法：python init_template.py [工作目录]
- 先在工作目录找 template.xls，找不到则使用 skill 内置模板
- template.xlsx 输出到工作目录
"""

import sys
import os
import xlrd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# skill 目录（本脚本所在目录）
SKILL_DIR = os.path.dirname(os.path.abspath(__file__))


def convert_xls_to_xlsx(xls_path, xlsx_path):
    """将 xls 的 3 个 sheet 复制到新的 xlsx 文件中。"""
    xls_book = xlrd.open_workbook(xls_path, formatting_info=True)
    xlsx_book = Workbook()
    xlsx_book.remove(xlsx_book.active)

    for sheet in xls_book.sheets():
        ws = xlsx_book.create_sheet(title=sheet.name)
        for r in range(sheet.nrows):
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                new_cell = ws.cell(row=r + 1, column=c + 1, value=cell.value)
                if cell.ctype == 3 and cell.value:
                    new_cell.number_format = 'YYYY-MM-DD HH:MM:SS'

    return xlsx_book


def add_mapping_sheets(xlsx_book):
    """添加 3 列分类映射 sheet 和 2 列账户映射 sheet。"""
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_align = Alignment(horizontal="center")

    # --- 分类映射（3 列：CSV交易分类, 一级分类, 二级分类）---
    # 首次使用时 CSV 列预填常见类目，一级/二级留空供用户（或 Agent）填写
    categories = [
        "餐饮美食", "日用百货", "医疗健康", "退款", "收入",
        "投资理财", "文化休闲", "交通出行", "数码电器",
        "家居家装", "生活服务", "商业服务", "爱车养车",
        "亲友代付", "服饰装扮", "住房物业", "运动户外",
        "充值缴费", "信用借还", "消费借贷", "转账红包", "其他",
    ]

    ws_cat = xlsx_book.create_sheet(title="分类映射")
    for col, title in enumerate(["CSV交易分类", "一级分类", "二级分类"], 1):
        cell = ws_cat.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, cat in enumerate(categories):
        ws_cat.cell(row=i + 2, column=1, value=cat)
    ws_cat.column_dimensions['A'].width = 16
    ws_cat.column_dimensions['B'].width = 14
    ws_cat.column_dimensions['C'].width = 14

    # --- 账户映射（2 列：CSV支付方式, 你的账户名）---
    payment_methods = [
        "账户余额", "余额宝", "账户余额(个人余额)",
        "招商银行储蓄卡(XXXX)", "招商银行信用卡(XXXX)",
        "账户余额&红包", "账户余额&碰友日立减",
        "账户余额&优惠", "账户余额&焕新折扣",
        "账户余额&企业码员工优惠",
        "支付宝小荷包(XXX)&红包",
        "(空)",
    ]

    ws_acc = xlsx_book.create_sheet(title="账户映射")
    for col, title in enumerate(["CSV支付方式", "你的账户名"], 1):
        cell = ws_acc.cell(row=1, column=col, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for i, method in enumerate(payment_methods):
        ws_acc.cell(row=i + 2, column=1, value=method)
    ws_acc.column_dimensions['A'].width = 30
    ws_acc.column_dimensions['B'].width = 20

    return xlsx_book


def main():
    work_dir = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()

    # 优先在工作目录找 template.xls，找不到则用 skill 内置模板
    xls_path = os.path.join(work_dir, "template.xls")
    if not os.path.exists(xls_path):
        xls_path = os.path.join(SKILL_DIR, "template.xls")
        if not os.path.exists(xls_path):
            print(f"错误：找不到 template.xls（已检查工作目录和 skill 内置目录）")
            sys.exit(1)

    xlsx_path = os.path.join(work_dir, "template.xlsx")

    print("转换 template.xls → template.xlsx ...")
    book = convert_xls_to_xlsx(xls_path, xlsx_path)

    print("添加分类映射 + 账户映射 sheet ...")
    book = add_mapping_sheets(book)

    # 把映射 sheet 排到最后
    sheet_order = ["支出", "收入", "转账", "分类映射", "账户映射"]
    for i, name in enumerate(sheet_order):
        if name in book.sheetnames:
            idx = book.sheetnames.index(name)
            book.move_sheet(name, offset=i - idx)

    book.save(xlsx_path)
    print(f"完成！已生成：{xlsx_path}")
    print("下一步：用 Excel 打开 template.xlsx，填写「分类映射」和「账户映射」sheet 中的空列，")
    print("        或直接运行 process_alipay.py，Agent 会自动推断未填写的映射。")


if __name__ == "__main__":
    main()
