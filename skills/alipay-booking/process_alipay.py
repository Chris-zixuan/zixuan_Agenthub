"""支付宝交易流水处理脚本。

读取支付宝 CSV 和用户模板，按映射规则生成记账用 Excel 文件。
用法：python process_alipay.py <工作目录>
"""

import csv
import os
import sys
import glob
from datetime import datetime
from openpyxl import load_workbook

# 忽略阈值：投资理财收益不高于此金额不记录
IGNORE_THRESHOLD = 0.04

# CSV 字段名（支付宝导出格式）
COL_TIME = "交易时间"
COL_CATEGORY = "交易分类"
COL_COUNTERPARTY = "交易对方"
COL_DESCRIPTION = "商品说明"
COL_PAY_TYPE = "收/支"
COL_AMOUNT = "金额"
COL_METHOD = "收/付款方式"
COL_STATUS = "交易状态"


def find_latest_csv(directory):
    """在目录中找最新的支付宝交易明细 CSV 文件。"""
    pattern = os.path.join(directory, "支付宝交易明细*.csv")
    files = glob.glob(pattern)
    if not files:
        raise FileNotFoundError(f"在 {directory} 中找不到支付宝交易明细 CSV 文件")
    return max(files, key=os.path.getmtime)


def parse_csv(csv_path):
    """解析支付宝 CSV，返回记录列表。

    每条记录为 dict，字段见 COL_* 常量定义。金额转换为 float。
    CSV 编码 GBK，前 24 行为说明/摘要（行 0-23），第 25 行（索引 23）为表头。
    """
    records = []
    with open(csv_path, encoding="gbk") as f:
        lines = f.readlines()

    if len(lines) < 25:
        print(f"警告：CSV 文件只有 {len(lines)} 行，可能不完整", file=sys.stderr)
        return records

    header_line = lines[23].strip().rstrip(",")
    headers = header_line.split(",")

    for line in lines[24:]:
        line = line.strip()
        if not line:
            continue

        # 支付宝 CSV 在订单号字段之间使用制表符，先归一化为逗号
        line = line.replace("\t", ",")
        parts = line.split(",")
        if len(parts) < len(headers):
            continue

        record = {}
        for i, h in enumerate(headers):
            record[h] = parts[i].strip() if i < len(parts) else ""

        try:
            amount_str = record.get(COL_AMOUNT, "0").replace(",", "").strip()
            record[COL_AMOUNT] = float(amount_str) if amount_str else 0.0
        except ValueError:
            record[COL_AMOUNT] = 0.0

        records.append(record)

    return records


def should_ignore(record):
    """判断记录是否应被忽略。"""
    status = record.get(COL_STATUS, "")
    category = record.get(COL_CATEGORY, "")
    pay_type = record.get(COL_PAY_TYPE, "")
    amount = record[COL_AMOUNT]

    if pay_type == "不计收支" and "投资理财" in category and amount <= IGNORE_THRESHOLD:
        return True

    if pay_type == "不计收支" and "关闭" in status:
        return True

    return False


def is_refund(record):
    """判断是否为退款（交易分类或状态含"退款"）。"""
    status = record.get(COL_STATUS, "")
    category = record.get(COL_CATEGORY, "")
    return "退款" in category or "退款" in status


def normalize_payment_method(method):
    """归一化支付方式——组合支付时取主账户。"""
    if not method:
        return "(空)"
    if "&" in method:
        return method.split("&")[0].strip()
    return method


def load_mappings(template_path):
    """从 template.xlsx 加载分类映射和账户映射。

    自动检测分类映射格式：
    - 3 列（CSV交易分类, 一级分类, 二级分类）→ 直接用作 category_map
    - 2 列（一级分类, 二级分类）→ 作为 taxonomy，需 Agent 推断 CSV→分类的映射

    返回: (category_map, account_map, taxonomy)
      category_map: {CSV交易分类: (一级分类, 二级分类)}  3列格式时填充
      account_map:  {CSV支付方式: 你的账户名}
      taxonomy:     [(一级分类, 二级分类), ...]  2列格式时填充
    """
    wb = load_workbook(template_path, data_only=True)
    category_map = {}
    account_map = {}
    taxonomy = []

    if "分类映射" in wb.sheetnames:
        ws = wb["分类映射"]
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        if not rows:
            print("警告：分类映射 sheet 为空", file=sys.stderr)
        else:
            header = rows[0]
            num_cols = sum(1 for h in header if h) if header else 0
            is_3col = num_cols >= 3

            for row in rows[1:]:
                if is_3col:
                    csv_cat = (row[0] or "").strip() if len(row) > 0 else ""
                    l1 = (row[1] or "").strip() if len(row) > 1 else ""
                    l2 = (row[2] or "").strip() if len(row) > 2 else ""
                    if csv_cat:
                        category_map[csv_cat] = (l1, l2)
                else:
                    l1 = (row[0] or "").strip() if len(row) > 0 else ""
                    l2 = (row[1] or "").strip() if len(row) > 1 else ""
                    if l1:
                        taxonomy.append((l1, l2))
    else:
        print("警告：template.xlsx 中缺少「分类映射」sheet，所有分类将留空", file=sys.stderr)

    if "账户映射" in wb.sheetnames:
        ws = wb["账户映射"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            csv_method = (row[0] or "").strip()
            user_account = (row[1] or "").strip()
            if csv_method:
                account_map[csv_method] = user_account or csv_method
    else:
        print("警告：template.xlsx 中缺少「账户映射」sheet，所有账户将使用原值", file=sys.stderr)

    wb.close()
    return category_map, account_map, taxonomy


def process_records(records, category_map, account_map):
    """将原始记录分类处理为 支出/收入/转账 三个列表。

    支出记录字段:(交易类型, 日期, 一级分类, 二级分类, 支付账户, 金额, 成员, 商家, 项目, 备注)
    收入记录字段:(交易类型, 日期, 一级分类, 二级分类, 收入账户, 金额, 成员, 商家, 项目, 备注)
    转账记录字段:(交易类型, 日期, 转出账户, 转入账户, 金额, 成员, 商家, 项目, 备注)
    """
    expenses = []
    incomes = []
    transfers = []

    for rec in records:
        if should_ignore(rec):
            continue

        pay_type = rec.get(COL_PAY_TYPE, "")
        csv_cat = rec.get(COL_CATEGORY, "")
        csv_method = rec.get(COL_METHOD, "")

        date_str = rec.get(COL_TIME, "")
        counterparty = rec.get(COL_COUNTERPARTY, "")
        description = rec.get(COL_DESCRIPTION, "")
        amount = rec[COL_AMOUNT]

        l1, l2 = category_map.get(csv_cat, ("", ""))
        normalized_method = normalize_payment_method(csv_method)
        account = account_map.get(normalized_method, normalized_method)
        note = f"{counterparty} {description}".strip()

        if pay_type == "支出" or (pay_type == "不计收支" and is_refund(rec)):
            # 退款金额取负
            final_amount = -amount if is_refund(rec) else amount
            expenses.append([
                csv_cat, date_str, l1, l2, account, final_amount,
                "", counterparty, "", note,
            ])

        elif pay_type == "收入":
            incomes.append([
                csv_cat, date_str, l1, l2, account, amount,
                "", counterparty, "", note,
            ])

        elif pay_type == "不计收支":
            # 不计收支的其余记录默认作为支出处理
            expenses.append([
                csv_cat, date_str, l1, l2, account, amount,
                "", counterparty, "", note,
            ])

    return expenses, incomes, transfers


def find_unmapped(records, category_map, account_map):
    """找出记录中存在但映射表中缺失的 CSV 分类和支付方式。

    返回:
      unmapped_cats:  {分类名: [示例交易列表], ...}，每笔示例含 counterparty/description
      unmapped_methods: [支付方式, ...]
    """
    csv_categories = {}
    csv_methods = set()

    for rec in records:
        cat = rec.get(COL_CATEGORY, "")
        method = normalize_payment_method(rec.get(COL_METHOD, ""))
        if cat:
            if cat not in csv_categories:
                csv_categories[cat] = []
            if len(csv_categories[cat]) < 5:
                csv_categories[cat].append({
                    "counterparty": rec.get(COL_COUNTERPARTY, ""),
                    "description": rec.get(COL_DESCRIPTION, ""),
                })
        if method:
            csv_methods.add(method)

    unmapped_cats = {k: v for k, v in sorted(csv_categories.items()) if k not in category_map}
    unmapped_methods = sorted(csv_methods - set(account_map.keys()))
    return unmapped_cats, unmapped_methods


def all_csv_categories_with_samples(records):
    """列出 CSV 中所有交易分类及示例，供 Agent 推断分类映射。"""
    cats = {}
    for rec in records:
        cat = rec.get(COL_CATEGORY, "")
        if cat and cat not in cats:
            cats[cat] = []
        if cat and len(cats[cat]) < 5:
            cats[cat].append({
                "counterparty": rec.get(COL_COUNTERPARTY, ""),
                "description": rec.get(COL_DESCRIPTION, ""),
            })
    return dict(sorted(cats.items()))


def append_mapping_rows(template_path, new_categories=None, new_accounts=None):
    """向 template.xlsx 的映射 sheet 追加新行。

    new_categories: [(csv交易分类, 一级分类, 二级分类), ...]
    new_accounts:   [(csv支付方式, 账户名), ...]
    """
    from openpyxl import load_workbook

    wb = load_workbook(template_path)

    if new_categories and "分类映射" in wb.sheetnames:
        ws = wb["分类映射"]
        for csv_cat, l1, l2 in new_categories:
            ws.append([csv_cat, l1, l2])
        print(f"已向分类映射追加 {len(new_categories)} 条规则")

    if new_accounts and "账户映射" in wb.sheetnames:
        ws = wb["账户映射"]
        for csv_method, acct in new_accounts:
            ws.append([csv_method, acct])
        print(f"已向账户映射追加 {len(new_accounts)} 条规则")

    wb.save(template_path)
    wb.close()


def fill_output(expenses, incomes, transfers, template_xlsx_path, output_path):
    """将处理好的数据填充到输出文件中。

    以 template.xlsx 的 支出/收入/转账 sheet 获取表头格式，重新写入数据。
    """
    wb = load_workbook(template_xlsx_path)

    def fill_sheet(ws, data, column_count):
        """清空 sheet 的示例数据（保留表头），写入新数据。"""
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for row_data in data:
            ws.append(row_data[:column_count])

    fill_sheet(wb["支出"], expenses, 10)
    fill_sheet(wb["收入"], incomes, 10)
    fill_sheet(wb["转账"], transfers, 9)

    # 删除映射 sheet（输出文件不需要这些配置）
    for s in ["分类映射", "账户映射"]:
        if s in wb.sheetnames:
            del wb[s]

    wb.save(output_path)
    wb.close()


def main():
    work_dir = sys.argv[1] if len(sys.argv) > 1 else os.getcwd()
    template_xlsx = os.path.join(work_dir, "template.xlsx")

    if not os.path.exists(template_xlsx):
        print(f"错误：找不到 template.xlsx，请先运行 init_template.py 初始化。")
        sys.exit(1)

    try:
        csv_path = find_latest_csv(work_dir)
        print(f"读取 CSV: {os.path.basename(csv_path)}")

        records = parse_csv(csv_path)
        print(f"解析 {len(records)} 条记录")

        category_map, account_map, taxonomy = load_mappings(template_xlsx)

        # 2 列分类映射 = 分类法模式，需 Agent 或用户完成 CSV→分类的映射
        if taxonomy and not category_map:
            print(f"\n检测到 2 列分类法（{len(taxonomy)} 个有效分类组合），需建立 CSV 交易分类到分类法的映射。")
            csv_cats = all_csv_categories_with_samples(records)
            print(f"CSV 中共 {len(csv_cats)} 个交易分类：\n")
            for cat, samples in csv_cats.items():
                sample_strs = [f"{s['counterparty']} - {s['description']}" for s in samples if s['counterparty'] or s['description']]
                detail = f"  示例: {'; '.join(sample_strs[:3])}" if sample_strs else ""
                print(f"  - {cat}{detail}")
            print(f"\n账户映射: {len(account_map)} 条规则")
            print("\n请根据以上示例推断每个 CSV 交易分类对应的 (一级分类, 二级分类)，")
            print("然后调用 build_csv_category_map() 写入映射，或手动在 template.xlsx 的「分类映射」sheet 中补充第三列。")
            sys.exit(0)

        print(f"分类映射: {len(category_map)} 条规则")
        print(f"账户映射: {len(account_map)} 条规则")

        expenses, incomes, transfers = process_records(records, category_map, account_map)
        print(f"支出: {len(expenses)} 条, 收入: {len(incomes)} 条, 转账: {len(transfers)} 条")

        # 检测映射表中缺失的类目和支付方式
        unmapped_cats, unmapped_methods = find_unmapped(records, category_map, account_map)
        if unmapped_cats:
            print(f"\n[!] 警告：{len(unmapped_cats)} 个交易分类在映射表中找不到：")
            for cat, samples in unmapped_cats.items():
                sample_strs = [f"{s['counterparty']} - {s['description']}" for s in samples if s['counterparty'] or s['description']]
                detail = f"  示例: {'; '.join(sample_strs[:3])}" if sample_strs else ""
                print(f"  - {cat}{detail}")
            print()
        if unmapped_methods:
            print(f"\n[!] 警告：{len(unmapped_methods)} 个支付方式在映射表中找不到：")
            for m in unmapped_methods:
                print(f"  - {m}")
            print()

        # 生成输出文件名（取第一条记录的日期，如果无记录则用当天日期）
        if records:
            date_prefix = records[0].get(COL_TIME, "")[:10].replace("-", "")
        else:
            date_prefix = datetime.now().strftime("%Y%m%d")
        output_path = os.path.join(work_dir, f"输出-{date_prefix}.xlsx")

        fill_output(expenses, incomes, transfers, template_xlsx, output_path)
        print(f"完成！输出：{output_path}")

    except FileNotFoundError as e:
        print(f"错误：{e}", file=sys.stderr)
        sys.exit(1)
    except KeyError as e:
        print(f"错误：template.xlsx 格式不正确，缺少 sheet: {e}", file=sys.stderr)
        sys.exit(1)
    except PermissionError as e:
        print(f"错误：文件被占用，请关闭 Excel 后重试: {e}", file=sys.stderr)
        sys.exit(1)
    except Exception as e:
        print(f"处理失败：{e}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
